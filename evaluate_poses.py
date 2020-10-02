#!/usr/bin/env python
from __future__ import print_function

import argparse
import copy
from evo.tools import log
log.configure_logging()
import numpy as np
import os
from tqdm import tqdm

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

import pandas as pd

from evo.tools import plot
from evo.tools.plot import PlotMode
from evo.tools.settings import SETTINGS
from evo.core.metrics import PoseRelation, Unit
from evo.core import metrics
from evo.core import sync
from evo.core import trajectory

# temporarily override some package settings
SETTINGS.plot_figsize = [6, 6]
SETTINGS.plot_split = True
SETTINGS.plot_usetex = False

from kitti_odometry import KittiEvalOdom

from evo.tools import file_interface
from evo.core import sync

import matplotlib.pyplot as plt

def function_statistics(type_of_stat, list_of_values):
    if type_of_stat == 'mean':
        return np.mean(list_of_values)
    if type_of_stat == 'median':
        return np.median(list_of_values)
    if type_of_stat == 'std':
        return np.std(list_of_values)
    if type_of_stat == 'var':
        return np.var(list_of_values)
    if type_of_stat == 'min':
        return np.min(list_of_values)
    if type_of_stat == 'max':
        return np.max(list_of_values)
    
# ---------------------------------- get_and_save_results_from_folder ----------------------------------------------------
def get_and_save_results_from_folder(folder_with_predicted_poses,category):
    
    global args
    global kitti_eval_tool
    global folder_with_gt_poses
    global output_folder
    global t
    global results
    
    values_for_excel = []
    columns_for_excel = []
    type_of_statistics = 'mean'
    for filename in sorted(os.listdir(folder_with_predicted_poses)):
        if not(os.path.exists(os.path.join(folder_with_gt_poses, filename))):
            print("file with gt poses doesn't exist for "+filename)
            continue
        if filename.find('.txt') == -1:
            continue
        seq_results = {}
        seq_results['name_seq'] = filename[:filename.rfind('.')]
        seq_results['category'] = category
        folder_name = seq_results['category']
        seq_results['metrics'] = {}
        seq_results['lost'] = False
        
        os.makedirs(os.path.join(output_folder, folder_name), exist_ok=True)
        output_folder_seq = os.path.join(output_folder, folder_name, filename[:filename.rfind('.')])
        os.makedirs(output_folder_seq, exist_ok=True)
        if os.path.isfile(os.path.join(output_folder, folder_name,"results.txt")):
            file_results_txt = open(os.path.join(output_folder, folder_name,"results.txt"), "a")
        else:
            file_results_txt = open(os.path.join(output_folder, folder_name,"results.txt"), "w")
            file_results_txt.write("translation_error(%) rotation_error(deg/m) ATE(m) APE_translation_error_median(m) APE_rotation_error_median(deg) dst_to_trgt\n")
        
        # -------------------------------------Getting results---------------------------------------------------
        if args.gt_format == 'kitti':        
            traj_ref = file_interface.read_kitti_poses_file(os.path.join(folder_with_gt_poses, filename))
        if args.gt_format == 'tum':        
            traj_ref = file_interface.read_tum_trajectory_file(os.path.join(folder_with_gt_poses, filename))
            seq_results["length_of_ref_traj"] = traj_ref.path_length
            end_time_gt = traj_ref.get_infos()["t_end (s)"]
        if args.gt_format == 'euroc':        
            traj_ref = file_interface.read_euroc_csv_trajectory(os.path.join(folder_with_gt_poses, filename))
        if args.result_format == 'kitti':
            traj_est = file_interface.read_kitti_poses_file(os.path.join(folder_with_predicted_poses, filename))
        if args.result_format == 'tum':
            traj_est = file_interface.read_tum_trajectory_file(os.path.join(folder_with_predicted_poses, filename))
            seq_results["length_of_estimated_traj"] = traj_est.path_length
        if args.result_format == 'euroc':
            traj_est = file_interface.read_euroc_csv_trajectory(os.path.join(folder_with_predicted_poses, filename))
        if args.result_format == 'tum' and args.gt_format == 'tum':
            seq_results["num_gt_poses"] = traj_ref.num_poses
            seq_results["num_predicted_poses"] = traj_est.num_poses
            traj_ref, traj_est = sync.associate_trajectories(traj_ref, traj_est, args.max_diff)
            end_time_est = traj_est.get_infos()["t_end (s)"]
            if (abs(end_time_est - end_time_gt) > 0.2) or (traj_est.get_infos()["t_start (s)"] > 0.2):
                print('LOST in track '+filename[:filename.rfind('.')])
                seq_results['lost'] = True
                results.append(seq_results)
                t.update(1)
                continue
        if args.alignment != None:
            traj_est = trajectory.align_trajectory(traj_est, traj_ref, correct_scale=args.alignment.find("scale") != -1, correct_only_scale=args.alignment=="scale")
        trajectory.align_trajectory_origin(traj_est, traj_ref)
        data = (traj_ref, traj_est)
        
        ape_metric_translation = metrics.APE(metrics.PoseRelation.translation_part)
        ape_metric_rotation = metrics.APE(metrics.PoseRelation.rotation_angle_deg)
        ape_metric_translation.process_data(data)
        ape_metric_rotation.process_data(data)
        ape_translation_statistics = ape_metric_translation.get_all_statistics()
        ape_rotation_statistics = ape_metric_rotation.get_all_statistics()
        
        ape_translation_statistics_plot = copy.deepcopy(ape_translation_statistics)
        ape_rotation_statistics_plot = copy.deepcopy(ape_rotation_statistics)
        ape_translation_statistics_plot.pop('sse')
        ape_translation_statistics_plot.pop('std')
        ape_translation_statistics_plot.pop('min')
        ape_translation_statistics_plot.pop('max')
        ape_rotation_statistics_plot.pop('sse')
        ape_rotation_statistics_plot.pop('std')
        ape_rotation_statistics_plot.pop('min')
        ape_rotation_statistics_plot.pop('max')
        
        kitti_trans_err, kitti_rot_err, ate = kitti_eval_tool.eval(traj_ref.poses_se3, 
                                                               traj_est.poses_se3, 
                                                               alignment=None)
    
        #---------------------------------adding results to variable seq_results for excel -----------------------------
        seq_results['metrics']['dist_to_trgt'] = traj_est.get_infos()['pos_end (m)'] - traj_ref.get_infos()['pos_end (m)']
        seq_results['metrics']['dist_to_trgt'] = np.sum(np.array(seq_results['metrics']['dist_to_trgt'])**2)**0.5
        seq_results['metrics']["Kitti trans err (%)"] = kitti_trans_err
        seq_results['metrics']["Kitti rot err (deg/m)"] = kitti_rot_err
        seq_results['metrics']["ATE (m)"] = ate
        seq_results['metrics']["APE(trans err) median (m)"] = ape_translation_statistics["median"]
        seq_results['metrics']["APE(rot err) median (deg)"] = ape_rotation_statistics["median"]
        #--------------------------------------------------------------------------------------------------------
        
        
        #-------------------------------------------------------------------------------------------------------    
    
        # --------------------------------printing results into console----------------------------------------------
        print('Results for "'+filename+'":')
        print('Kitti average translational error (%): {:.7f}'.format(kitti_trans_err))
        print('Kitti average rotational error (deg/m): {:.7f}'.format(kitti_rot_err))
        print('ATE (m): {:.7f}'.format(ate))
        print('APE(translation error) median (m): {:.7f}'.format(ape_translation_statistics["median"]))
        print('APE(rotation error) median (deg): {:.7f}'.format(ape_rotation_statistics["median"]))
        print('distance to target on the last frame: {:.7f}'.format(seq_results['metrics']['dist_to_trgt']))
        #------------------------------------------------------------------------------------------------------------
        
        #---------------------------------Saving results into overall results text file------------------------------
        file_results_txt.write('{:<24} '.format(filename[:filename.rfind('.')]))
        file_results_txt.write('{:>7.4f} '.format(kitti_trans_err))
        file_results_txt.write('{:>7.4f} '.format(kitti_rot_err))
        file_results_txt.write('{:>7.4f} '.format(ate))
        file_results_txt.write('{:>7.4f} '.format(ape_translation_statistics["median"]))
        file_results_txt.write('{:>7.4f} '.format(ape_rotation_statistics["median"]))
        file_results_txt.write('{:>7.4f}\n'.format(seq_results['metrics']['dist_to_trgt']))
        #------------------------------------------------------------------------------------------------------------
    
        # --------------------------------Saving metrics to text file for one track----------------------------------
        txt_filename = filename[:filename.rfind('.')]+"_metrics.txt"
        with open(os.path.join(output_folder_seq, txt_filename), "w") as txt_file:
            txt_file.write('Kitti average translational error (%): {:.7f}\n'.format(kitti_trans_err))
            txt_file.write('Kitti average rotational error (deg/m): {:.7f}\n'.format(kitti_rot_err))
            txt_file.write('ATE (m): {:.7f}\n'.format(ate))
            txt_file.write('APE(translation error) median (m): {:.7f}\n'.format(ape_translation_statistics["median"]))
            txt_file.write('APE(rotation error) median (deg): {:.7f}\n'.format(ape_rotation_statistics["median"]))
            txt_file.write('Distance to target on the last frame: {:.7f}\n'.format(seq_results['metrics']['dist_to_trgt']))
        #---------------------------------------------------------------------------------------------------------
    
        # ---------------------------------Saving values of errors for each frame to text file------------------------
        # ------------------------------------------for translation errors----------------------------------------
        txt_filename = filename[:filename.rfind('.')]+"_APE(translation)_errors.txt"
        output_folder_seq_translation = os.path.join(output_folder_seq,"translation")
        output_folder_seq_rotation = os.path.join(output_folder_seq,"rotation")
        os.makedirs(output_folder_seq_translation, exist_ok=True)
        os.makedirs(output_folder_seq_rotation, exist_ok=True)
        with open(os.path.join(output_folder_seq_translation, txt_filename), "w") as txt_file:
            for error in ape_metric_translation.error:
                txt_file.write('{:.10f}\n'.format(error))
        # -----------------------------------------for rotation degree errors--------------------------------------
        txt_filename = filename[:filename.rfind('.')]+"_APE(rotation_deg)_errors.txt"
        with open(os.path.join(output_folder_seq_rotation, txt_filename), "w") as txt_file:
            for error in ape_metric_rotation.error:
                txt_file.write('{:.10f}\n'.format(error))
        #----------------------------------------------------------------------------------------------------------
            
        # ---------------------------------------Saving plot of errors of each frame------------------------------
        # ------------------------------------------for translation errors----------------------------------------
        plot_collection = plot.PlotCollection("Example")
        fig_1 = plt.figure(figsize=(8, 8))
        plot.error_array(fig_1, ape_metric_translation.error, 
                         name="APE", title=str(ape_metric_translation), xlabel="Index of frame", ylabel='Error')
        plot_collection.add_figure("raw", fig_1)
        plot_filename = filename[:filename.rfind('.')]+"_APE(translation)_errors.png"
        plt.savefig(os.path.join(output_folder_seq_translation, plot_filename))
        plt.close(fig_1)
        # -----------------------------------------for rotation degree errors--------------------------------------
        plot_collection = plot.PlotCollection("Example")
        fig_1 = plt.figure(figsize=(8, 8))
        plot.error_array(fig_1, ape_metric_rotation.error, 
                         name="APE", title=str(ape_metric_rotation), xlabel="Index of frame", ylabel='Error')
        plot_collection.add_figure("raw", fig_1)
        plot_filename = filename[:filename.rfind('.')]+"_APE(rotation)_errors.png"
        plt.savefig(os.path.join(output_folder_seq_rotation,plot_filename))
        plt.close(fig_1)
        #-----------------------------------------------------------------------------------------------------------
    
        # -----------------------------------------Saving trajectory plot------------------------------------------- 
        # ------------------------------------------for translation errors----------------------------------------
        fig_2 = plt.figure(figsize=(8, 8))
        ax = plot.prepare_axis(fig_2, plot_mode)
        plot.traj(ax, plot_mode, traj_ref, '--', 'gray', 'reference')
        plot.traj_colormap( ax, traj_est, ape_metric_translation.error, plot_mode, 
                           min_map=ape_translation_statistics["min"],
                           max_map=ape_translation_statistics["max"], title="APE translation mapped onto trajectory")
        plot_collection.add_figure("traj (error)", fig_2)
        plot_filename = filename[:filename.rfind('.')]+"_APE(translation)_map.png"
        plt.savefig(os.path.join(output_folder_seq_translation,plot_filename))
        plt.close(fig_2)
        # -----------------------------------------for rotation degree errors--------------------------------------
        fig_2 = plt.figure(figsize=(8, 8))
        ax = plot.prepare_axis(fig_2, plot_mode)
        plot.traj(ax, plot_mode, traj_ref, '--', 'gray', 'reference')
        plot.traj_colormap( ax, traj_est, ape_metric_rotation.error, plot_mode, 
                           min_map=ape_rotation_statistics["min"],
                           max_map=ape_rotation_statistics["max"], title="APE rotation mapped onto trajectory")
        plot_collection.add_figure("traj (error)", fig_2)
        plot_filename = filename[:filename.rfind('.')]+"_APE(rotation)_map.png"
        plt.savefig(os.path.join(output_folder_seq_rotation,plot_filename))
        plt.close(fig_2)
        #-----------------------------------------------------------------------------------------------------------
        print()
        
        active_worksheet = wb['sheet1']
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="thick", color="000000")
        medium = Side(border_style="medium", color="000000")
        font_header = Font(name='Arial',
                       size=10,
                       bold=True,
                       italic=False,
                       vertAlign=None,
                       underline='none',
                       strike=False,
                       color='FF000000')
        font_values = Font(name='Arial',
                       size=10,
                       bold=False,
                       italic=False,
                       vertAlign=None,
                               underline='none',
                       strike=False,
                       color='FF000000')

        active_worksheet.row_dimensions[2].height = 35
        
        file_results_txt.close()
        results.append(seq_results)
        t.update(1)
        

def output_summary(results):
    global output_folder
    
    statistics = {}
    
    for seq_results in results:
        folder_name = seq_results["category"]
        if not(folder_name in statistics.keys()):
            statistics[folder_name] = {
                "Kitti trans err (%)": [],
                "Kitti rot err (deg/m)": [],
                "ATE (m)": [],
                "APE(trans err) median (m)": [],
                "APE(rot err) median (deg)": [],
                "dist_to_trgt": [],
                "lost": int(seq_results['lost'])
            }
            
        else:
            statistics[folder_name]["lost"] = statistics[folder_name]["lost"] + int(seq_results['lost'])
    
    for seq_results in results:
        if seq_results["lost"]==True:
            continue
        folder_name = seq_results["category"]
        statistics[folder_name]["Kitti trans err (%)"].append(seq_results['metrics']["Kitti trans err (%)"])
        statistics[folder_name]["Kitti rot err (deg/m)"].append(seq_results['metrics']["Kitti rot err (deg/m)"])
        statistics[folder_name]["ATE (m)"].append(seq_results['metrics']["ATE (m)"])
        statistics[folder_name]["APE(trans err) median (m)"].append(seq_results['metrics']["APE(trans err) median (m)"])
        statistics[folder_name]["APE(rot err) median (deg)"].append(seq_results['metrics']["APE(rot err) median (deg)"])
        statistics[folder_name]["dist_to_trgt"].append(seq_results['metrics']["dist_to_trgt"])
            
    for folder_name in statistics.keys():
        file_results_txt = open(os.path.join(output_folder, folder_name,"results.txt"), "a")
        file_results_txt.write("\n")
        file_results_txt.write('num lost, %: {}\n'.format(100*statistics[folder_name]['lost']/len(results)))
        if 100*statistics[folder_name]['lost']/len(results) == 100:
            continue
        file_results_txt.write("Mean:\n")
        file_results_txt.write('{:.4f} {:.4f} {:.4f} {:.4f} {:.4f} {:.4f}\n'.\
                               format(np.mean(np.array(statistics[folder_name]["Kitti trans err (%)"])),\
                                     np.mean(np.array(statistics[folder_name]["Kitti rot err (deg/m)"])),\
                                     np.mean(np.array(statistics[folder_name]["ATE (m)"])),\
                                     np.mean(np.array(statistics[folder_name]["APE(trans err) median (m)"])),\
                                     np.mean(np.array(statistics[folder_name]["APE(rot err) median (deg)"])),\
                                     np.mean(np.array(statistics[folder_name]["dist_to_trgt"]))))
        file_results_txt.write("Median:\n")
        file_results_txt.write('{:.4f} {:.4f} {:.4f} {:.4f} {:.4f} {:.4f}\n'.\
                               format(np.median(np.array(statistics[folder_name]["Kitti trans err (%)"])),\
                                     np.median(np.array(statistics[folder_name]["Kitti rot err (deg/m)"])),\
                                     np.median(np.array(statistics[folder_name]["ATE (m)"])),\
                                     np.median(np.array(statistics[folder_name]["APE(trans err) median (m)"])),\
                                     np.median(np.array(statistics[folder_name]["APE(rot err) median (deg)"])),\
                                     np.median(np.array(statistics[folder_name]["dist_to_trgt"]))))
        file_results_txt.write("Min:\n")
        file_results_txt.write('{:.4f} {:.4f} {:.4f} {:.4f} {:.4f} {:.4f}\n'.\
                               format(np.min(np.array(statistics[folder_name]["Kitti trans err (%)"])),\
                                     np.min(np.array(statistics[folder_name]["Kitti rot err (deg/m)"])),\
                                     np.min(np.array(statistics[folder_name]["ATE (m)"])),\
                                     np.min(np.array(statistics[folder_name]["APE(trans err) median (m)"])),\
                                     np.min(np.array(statistics[folder_name]["APE(rot err) median (deg)"])),\
                                     np.min(np.array(statistics[folder_name]["dist_to_trgt"]))))
        file_results_txt.write("Max:\n")
        file_results_txt.write('{:.4f} {:.4f} {:.4f} {:.4f} {:.4f} {:.4f}\n'.\
                               format(np.max(np.array(statistics[folder_name]["Kitti trans err (%)"])),\
                                     np.max(np.array(statistics[folder_name]["Kitti rot err (deg/m)"])),\
                                     np.max(np.array(statistics[folder_name]["ATE (m)"])),\
                                     np.max(np.array(statistics[folder_name]["APE(trans err) median (m)"])),\
                                     np.max(np.array(statistics[folder_name]["APE(rot err) median (deg)"])),
                                     np.max(np.array(statistics[folder_name]["dist_to_trgt"]))))
        file_results_txt.write("Variance:\n")
        file_results_txt.write('{:.4f} {:.4f} {:.4f} {:.4f} {:.4f} {:.4f}\n'.\
                               format(np.var(np.array(statistics[folder_name]["Kitti trans err (%)"])),\
                                     np.var(np.array(statistics[folder_name]["Kitti rot err (deg/m)"])),\
                                     np.var(np.array(statistics[folder_name]["ATE (m)"])),\
                                     np.var(np.array(statistics[folder_name]["APE(trans err) median (m)"])),\
                                     np.var(np.array(statistics[folder_name]["APE(rot err) median (deg)"])),
                                     np.var(np.array(statistics[folder_name]["dist_to_trgt"]))))
        file_results_txt.close()
    
#     active_worksheet = wb['sheet1']
#     thin = Side(border_style="thin", color="000000")
#     thick = Side(border_style="thick", color="000000")
#     medium = Side(border_style="medium", color="000000")
    
#     font_header = Font(name='Arial',
#                        size=10,
#                        bold=True,
#                        italic=False,
#                        vertAlign=None,
#                        underline='none',
#                        strike=False,
#                        color='FF000000')
#     font_values = Font(name='Arial',
#                        size=10,
#                        bold=False,
#                        italic=False,
#                        vertAlign=None,
#                        underline='none',
#                        strike=False,
#                        color='FF000000')

#     df = pd.DataFrame.from_dict({"results":values_for_excel}, columns=columns_for_excel, orient='index')
#     active_worksheet.row_dimensions[2].height = 35
        
#     #--------------------------------------------- output statistics ------------------------------------------------
#     shift = shift_statistics
#     active_cell = active_worksheet.cell(row = 1, column=shift+1)
#     active_worksheet.merge_cells(start_row=1, start_column=shift+1, end_row=1, end_column=shift+5)
#     active_cell.border = Border(top=thick, left=thick, right=thick, bottom=thin)
#     active_cell.value = type_of_statistics + ' errors'
#     active_cell.font = font_header
#     active_cell.alignment = Alignment(horizontal='center', vertical='center')

#     for i in range(5):
#         active_worksheet.column_dimensions[get_column_letter(shift+i+1)].width = 11
#         active_cell = active_worksheet.cell(row=2, column=shift+i+1)
#         active_cell.value = columns_for_excel[i]
#         active_cell.font = font_header
#         active_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, wrapText=True)
#         active_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
#         if i == 4:
#             active_cell.border = Border(top=thin, left=thin, right=thick, bottom=thin)

#         for i in range(5):
#             active_cell = active_worksheet.cell(row=3+shift_dataset, column=shift+i+1)
#             active_cell.value = '{:.5f}'.format(function_statistics(type_of_statistics, map_of_number_and_error[i])).replace(".", ",")
#             active_cell.font = font_values
#             active_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
#             if i == 4:
#                 active_cell.border = Border(top=thin, left=thin, right=thick, bottom=thin)
#                 active_cell.alignment = Alignment(horizontal='left', vertical='top')
#         active_cell = active_worksheet.cell(row=3+shift_dataset, column=1)
#         active_cell.value = folder_name
#         active_cell.font = font_header
#         #-------------------------------------------------------------------------------------------------------------

#         #-------------------- output names of sequences, names of errors and values into Excel -----------------------
#         shift = shift_statistics+shift_values
#         for i, filename in enumerate(sorted(os.listdir(folder_with_predicted_poses))): # iterate over 
#             target_col = len(active_worksheet[1])
#             seq_found = False
#             for col in active_worksheet.iter_cols(min_row=1, max_row=1):
#                 for cell in col:
#                     if str(cell.value) == filename[:filename.rfind('.')]:
#                         target_col = cell.col_idx
#                         seq_found = True
#             if seq_found:
#                 #-------- output values and names of errors into Excel ------------------------
#                 active_cell = active_worksheet.cell(row = 3+shift_dataset, column=target_col)
#                 active_worksheet.column_dimensions[get_column_letter(target_col)].width = 10
#                 active_cell.value = '{}'.format(df.iloc[0,i]).replace(".", ",")
#             else:
#                 #-------------------- output names of sequences ------------------------------
#                 active_cell = active_worksheet.cell(row = 1, column=target_col+5)
#                 active_cell.alignment = Alignment(horizontal='center', vertical='center')
#                 active_cell.value = filename[:filename.rfind('.')]
#                 active_cell.border = Border(top=thin, left=medium, right=medium, bottom=thin)
#                 active_cell.font = font_header
#                  #-------- output names of errors and values errors into Excel ------------------------
#                 for j, key in enumerate(results[folder_name]):
#                     active_cell = active_worksheet.cell(row = 3+shift_dataset, column=target_col+j)
#                     active_cell.value = '{}'.format(results[folder_name][key]).replace(".", ",")
#                         active_worksheet.merge_cells(start_row=1, 
#                                      start_column=target_col+1, 
#                                      end_row=1, 
#                                      end_column=target_col+1+4)   
        #-------- output values and names of errors into Excel ----------------------------------
#         for i in range(len(df.iloc[0,:])):
#             active_cell.value = '{}'.format(df.iloc[0,i]).replace(".", ",")
#             active_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
#             active_cell.font = font_values
#             active_worksheet.cell(row = 2, column=target_col).value = columns_for_excel[i%5]
#             active_worksheet.cell(row = 2, column=target_col).font = font_values
#             active_worksheet.cell(row = 2, column=target_col).border = Border(top=thin, 
#                                                                               left=thin, 
#                                                                               right=thin, 
#                                                                               bottom=thin)
#             active_worksheet.cell(row = 2, column=target_col).alignment = Alignment(horizontal='left', 
#                                                                                     vertical='top', 
#                                                                                     wrap_text=True, 
#                                                                                     wrapText=True)
#             if i % 5 == 4:
#                 active_cell.border = Border(top=thin, left=thin, right=medium, bottom=thin)
#                 active_worksheet.cell(row = 2, column=target_col).border = Border(top=thin, 
#                                                                                   left=thin, 
#                                                                                   right=medium, 
#                                                                                   bottom=thin)
                    #--------------------------------------------------------------------
        
#         shift_dataset = shift_dataset +1
# -----------------------------------------------------------------------------------------------------------------------
    # Saving results to excel file
#     wb.save(os.path.join(output_folder,"results.xlsx"))
    

def main():
    
    global folder_with_gt_poses
    global wb
    global args
    global kitti_eval_tool
    global output_folder
    global plot_mode
    global t
    global results
    
    argparser = argparse.ArgumentParser(description='Evaluation of poses')
    argparser.add_argument('--dir_gt', help='directory with gt poses', 
                           default='/media/cds-s/data2/Datasets/Husky-NKBVS/gt_poses_camera(with_cam_lidar_timestamps)')
    
    argparser.add_argument('--dir_result', help='directory with predicted poses', default='/media/cds-s/data2/Datasets/Husky-NKBVS/OpenVSLAM_results')
    
    argparser.add_argument('--dir_output', help='output directory, where results will be', 
                           default='/media/cds-s/data2/Datasets/Husky-NKBVS/result_evaluation')
    
    argparser.add_argument('--gt_format', choices=["kitti", "tum", "euroc"], help='format of gt poses: "kitti" or "tum" or "euroc"', default='kitti', required=False)

    argparser.add_argument('--result_format', choices=["kitti", "tum", "euroc"], help='format of result poses: "kitti" or "tum" or "euroc"', default='kitti', required=False)
    
    argparser.add_argument('--projection', choices=["xy", "yx", "xz", "zx", "yz", "zy"],
                           help='projection on which trajectory will plotted. Possible variants: "xy", "yx", "xz", "zx", "yz", "zy"', default = "xz", type=str)
    
    argparser.add_argument('--max_diff', help="maximum difference between timestamps in case when gt foramt is tum and result format is tum. By default it's 1/(2*FPS)=0.05 for FPS=10", 
                           default=10, type=float)

    argparser.add_argument('--alignment', choices=["scale", "6dof", "7dof", "scale_7dof"], help="Type of alignment. Choices are: 'scale', '6dof', '7dof', 'scale_7dof'", 
                           default=None, type=str)
    
    args = argparser.parse_args()
    folder_with_gt_poses = args.dir_gt
    folder_with_predicted_poses = args.dir_result
    output_folder = args.dir_output
    plot_mode = PlotMode(args.projection)
    
    plt.ioff()
    kitti_eval_tool = KittiEvalOdom()
    os.makedirs(output_folder, exist_ok=True)
    
    wb = Workbook()
    for sheet_name in wb.sheetnames:
        del wb[sheet_name]
    sheet1 = wb.create_sheet('sheet1',0)
    
    results = []
    proccessed_files_in_root_res_dir = False
    noOfFiles = 0
    for base, dirs, files in os.walk(folder_with_predicted_poses):
        for Files in files:
            noOfFiles += 1
    t = tqdm(total=noOfFiles)
    for filename in sorted(os.listdir(folder_with_predicted_poses)):
        if os.path.isfile(os.path.join(folder_with_predicted_poses, filename)) and not(proccessed_files_in_root_res_dir):
            category = folder_with_predicted_poses.rstrip('/')
            category = category.split('/')[-1]
            get_and_save_results_from_folder(folder_with_predicted_poses, category)
            proccessed_files_in_root_res_dir = True
            output_summary(results)
            results = []
        else:
            if filename.find('.txt') != -1:
                continue
            category = filename.rstrip('/')
            get_and_save_results_from_folder(os.path.join(folder_with_predicted_poses, filename), category)
            output_summary(results)
            results = []
    t.close()
    for seq_results in results:
        if seq_results['metrics'] == {}:
            continue
        if seq_results["metrics"]["Kitti trans err (%)"] == 0:
            print("'Kitti trans err (%)' = 0")
            print("dataset - "+seq_results['name'])
            print()
        if seq_results["num_gt_poses"]*0.5 > seq_results["num_predicted_poses"]:
            print("few predicted poses:")
            print("num predicted poses "+str(seq_results["num_predicted_poses"])+\
                  ", num gt poses "+str(seq_results["num_gt_poses"]))
            print("dataset - "+seq_results['name'])
            print()
    
if __name__ == '__main__':
    main()
